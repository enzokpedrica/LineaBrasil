import openpyxl
from openpyxl.styles import Alignment, Font
import subprocess
import pandas as pd

file = "C:/git/linea-ordenacao-requisicao/Lista_Map.CSV"


df= pd.read_csv(file, delimiter=';', encoding='latin1')


wb = openpyxl.load_workbook('layout.xlsx')
ws = wb.active
ws.title = "Tabela"

# Função para ajustar o tamanho da fonte com base no comprimento do texto
def ajustar_tamanho_fonte(texto, tamanho_base=150):
    comprimento = len(texto)
    if comprimento <= 5:
        return tamanho_base
    elif comprimento <= 6:
        return tamanho_base - 20
    else:
        return tamanho_base - 35

# Definir a quantidade
quantidade = input("Digita a Quantidade: ")
ws["B3"] = quantidade
cell = ws["B3"]
tamanho_fonte = ajustar_tamanho_fonte(quantidade, tamanho_base=150)
cell.font = Font(size=tamanho_fonte)
cell.alignment = Alignment(wrapText=True, horizontal='center', vertical='center')

# Solicitar o código do produto ao usuário
codigo_produto_int = int(input("Digite o Código do Produto: "))
codigo_produto_str = str(codigo_produto_int)

# Buscar a descrição do produto no DataFrame
produto = df.loc[df['Cod Produto'] == codigo_produto_int, 'Produto'].values[0]

# Definir o produto
ws["B2"] = produto
cell = ws["B2"]
cell.font = Font(size=60)
cell.alignment = Alignment(wrapText=True, horizontal='center', vertical='center')

# #--Numero Pedido
ws["B4"] = codigo_produto_str[-4:]
cell = ws["B4"]
cell.font = Font(size=190)
cell.alignment = Alignment(wrapText=True,horizontal='center', vertical='center')


# #--Volume
# ws["B5"] = "10"
# cell = ws["B5"]
# cell.font = Font(size=110)
# cell.alignment = Alignment(wrapText=True,horizontal='center', vertical='center')

for cell in ws[1]:
    cell.alignment = Alignment(horizontal='center')

# Salve o arquivo Excel
wb.save('tabela_imprimir.xlsx')

# Abra o arquivo Excel salvo
subprocess.Popen(['start', 'excel.exe', 'tabela_imprimir.xlsx'], shell=True)

print("Arquivo Excel aberto com sucesso!")