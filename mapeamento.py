import pandas as pd
import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import Border, Side
import subprocess

# Carregar o arquivo Excel
file = "Planilha.xlsx"
wb = load_workbook(file)
sheet = wb.active

# Definir borda padrão para remover qualquer borda anterior
no_border = Border(left=Side(style=None), right=Side(style=None), top=Side(style=None), bottom=Side(style=None))

# Iterar por todas as células e remover bordas
for row in sheet.iter_rows():
    for cell in row:
        cell.border = no_border  # Remover bordas

# Salvar a planilha com as bordas removidas
wb.save("Planilha_para_manipulacao.xlsx")

file = "Planilha.xlsx"
df = pd.read_excel(file, engine = "openpyxl")

print(df[['Produto','Descrição Produto','Qtde Mov']])

file_map = "Lista_Map.CSV"
df_map = pd.read_csv(file_map, delimiter=';', encoding='latin1')

df_merged = pd.merge(df, df_map, left_on='Produto', right_on='Cod Produto', how="inner")

# Ordenar e selecionar as colunas necessárias
df_merged = df_merged.sort_values(by='Secao')
df_merged = df_merged[['Produto_x', 'Descrição Produto', 'Qtde Mov', 'Rua', 'Secao', 'Andar']]
df_merged = df_merged.drop_duplicates(subset=['Descrição Produto'])

df_merged.to_excel('mapeamento_pre_arquivo.xlsx', index=False, engine='openpyxl')

wb = openpyxl.load_workbook('mapeamento_pre_arquivo.xlsx')
ws = wb.active
ws.title = "Tabela"

# Salve o arquivo Excel
wb.save('mapeamento.xlsx')

# Abra o arquivo Excel salvo
subprocess.Popen(['start', 'excel.exe', 'mapeamento.xlsx'], shell=True)

