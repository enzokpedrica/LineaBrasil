import pandas as pd
import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import Border, Side, Font, Alignment
import subprocess

def mapeamentoItens():
    # Declarando variáveis para os caminhos
    arquivo_exportado = r"C:\git\linea-ordenacao-requisicao\Mapeamento\Planilha.xlsx"
    arquivo_posicao = r"C:\git\linea-ordenacao-requisicao\Lista_Map.CSV"
    arquivo_mapeado = r'C:\git\linea-ordenacao-requisicao\Mapeamento\mapeamento.xlsx'

    #-------------------------------------------
    # Merge entre Planilhas do Senior com o Mapa
    #-------------------------------------------
    file = arquivo_exportado
    df = pd.read_excel(file, engine = "openpyxl")

    file_map = arquivo_posicao
    df_map = pd.read_csv(file_map, delimiter=';', encoding='latin1')

    df_merged = pd.merge(df, df_map, left_on='Produto', right_on='Cod Produto', how="inner")

    # Ordenar e selecionar as colunas necessárias
    df_merged = df_merged.sort_values(by='Secao')
    df_merged = df_merged[['Produto_x', 'Descrição Produto', 'Qtde Mov', 'Rua', 'Secao', 'Andar']]
    df_merged = df_merged.drop_duplicates(subset=['Descrição Produto'])
    df_merged = df_merged.rename(columns={'Produto_x' : "Codigo", 'Descrição Produto' : "Descrição", 'Qtde Mov' : "Quantidade" })

    df_merged.to_excel(arquivo_mapeado, index=False, engine='openpyxl')

    #-------------------------------------------
    # Organização "estética" do Excel
    #-------------------------------------------
    wb = openpyxl.load_workbook(arquivo_mapeado)
    ws = wb.active
    ws.title = "Tabela"

    # Definindo o tamanho das colunas
    column_widths = {
        'A': 12,  # Produto_x
        'B': 50,  # Descrição Produto
        'C': 15,  # Qtde Mov
        'D': 7,  # Rua
        'E': 7,  # Secao
        'F': 7,  # Andar
    }

    for coluna, tamanho in column_widths.items():
        ws.column_dimensions[coluna].width = tamanho

    # Define a fonte com tamanho de 12
    fonte_padrao = Font(size=12)

    # Aplica a fonte para todas as células
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.font = fonte_padrao

    title_font = Font(size=12, bold=True)

    for cell in ws[1]:
        cell.alignment = Alignment(horizontal='center')

    # Ajustando posição dos valores nas células
    for row in ws[2:ws.max_row]:
        for col in [3, 4, 5, 6]:
            cell = row[col - 1]
            cell.alignment = Alignment(horizontal='center')
        for col in [1]:
            cell = row[col - 1]
            cell.alignment = Alignment(horizontal='left')

    # Adicione bordas de separação para linhas e colunas
    borda = Border(left=Side(style='thin'), 
                        right=Side(style='thin'), 
                        top=Side(style='thin'), 
                        bottom=Side(style='thin'))

    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.border = borda

    # Salve o arquivo Excel
    wb.save(arquivo_mapeado)

    # Abra o arquivo Excel salvo
    subprocess.Popen(['start', 'excel.exe', arquivo_mapeado], shell=True)
    print("Arquivo Excel aberto com sucesso!")