import openpyxl
from openpyxl.styles import Alignment, Font
import subprocess
import pandas as pd

file = "C:/git/linea-ordenacao-requisicao/Lista_Map.CSV"


df= pd.read_csv(file, delimiter=';', encoding='latin1')


wb = openpyxl.load_workbook('layout_fita.xlsx')
ws = wb.active
ws.title = "Tabela"

# Função para ajustar o tamanho da fonte com base no comprimento do texto
def ajustar_tamanho_fonte(texto, tamanho_base=150):
    comprimento = len(texto)
    if comprimento <= 5:
        return tamanho_base
    elif comprimento <= 6:
        return tamanho_base - 20
    else:
        return tamanho_base - 35

# Definir a quantidade
quantidade = input("Digita a Quantidade: ")
ws["B4"] = quantidade
cell = ws["B4"]
tamanho_fonte = ajustar_tamanho_fonte(quantidade, tamanho_base=150)
cell.font = Font(size=tamanho_fonte)
cell.alignment = Alignment(wrapText=True, horizontal='center', vertical='center')

# Solicitar o código do produto ao usuário
codigo_produto_int = int(input("Digite o Código do Produto: "))
codigo_produto_str = str(codigo_produto_int)

# Buscar a descrição do produto no DataFrame
produto = df.loc[df['Cod Produto'] == codigo_produto_int, 'Produto'].values[0]

# Definir o produto
if codigo_produto_int == 150001004:
    produto = "Fita Borda Papel 18MM"
elif codigo_produto_int == 150001005:
    produto = "Fita Borda Papel 28MM"
elif codigo_produto_int == 150001006:
    produto = "Fita Borda Papel 40MM"   
else:
    produto = "Fita Borda PS"

ws["B2"] = produto
cell = ws["B2"]
cell.font = Font(size=50)
cell.alignment = Alignment(wrapText=True, horizontal='center', vertical='center')

#Definir a Cor
cores = pd.read_csv("Cores.CSV")
cores['ID'] = range(1, len(cores) + 1)
cores.set_index('ID', inplace=True)
print(cores)
cor_escolhida = int(input("Qual a cor? "))
cor_selecionada = cores.loc[cor_escolhida, 'Tipo/Cor']
ws["B3"] = cor_selecionada
cell = ws["B3"]
cell.font = Font(size=60)
cell.alignment = Alignment(wrapText=True,horizontal='center', vertical='center')

# #--Numero Pedido
ws["B5"] = codigo_produto_str
cell = ws["B5"]
cell.font = Font(size=90)
cell.alignment = Alignment(wrapText=True,horizontal='center', vertical='center')


for cell in ws[1]:
    cell.alignment = Alignment(horizontal='center')

# Salve o arquivo Excel
wb.save('tabela_imprimir.xlsx')

# Abra o arquivo Excel salvo
subprocess.Popen(['start', 'excel.exe', 'tabela_imprimir.xlsx'], shell=True)

print("Arquivo Excel aberto com sucesso!")



