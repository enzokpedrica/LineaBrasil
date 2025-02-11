import pandas as pd
import re
import openpyxl
from openpyxl.styles import Border, Side, Alignment
import subprocess

file = "teste.TXT"

colspecs = [(3, 12), (12, 16), (32, 80), (80, 125)]
colnames = ['Codigo', 'Deriv', 'Descricao', 'Quantidade']

df = pd.read_fwf(file, colspecs=colspecs, skiprows=9, header=None, names=colnames, encoding='latin1')

# Função para limpar a coluna de Quantidade
def clean_quantity(qtd):
    if isinstance(qtd, str):
        return re.sub(r'[^0-9,]', '', qtd.strip()).replace(',', '.')
    return qtd

# Aplica a função de limpeza na coluna 'Quantidade'
df['Quantidade'] = df['Quantidade'].apply(clean_quantity)
df['Quantidade'] = df['Quantidade'].str.strip()
df['Quantidade'] = pd.to_numeric(df['Quantidade'])


df = df.dropna(subset=['Quantidade'])
df["Deriv"] = df["Deriv"].fillna("-")
df['Descricao'] = df['Descricao'].str.slice(0, 30).str.strip()
df['Quantidade'] = df['Quantidade'].apply(lambda x: f"{x:,.2f}".replace(',', 'v').replace('.', ',').replace('v', '.'))
df['Codigo'] = df['Codigo'].astype(int)

#--- LEITURA DO ARQUIVO NOVO ---
file_map = "Lista_Map.CSV"
df_map = pd.read_csv(file_map, delimiter=';', encoding='latin1')

# MERGE
df_merged = pd.merge(df, df_map, left_on='Codigo', right_on='Cod Produto', how='left')
df_merged.drop(columns=['Cod Produto', 'Produto', 'Deriv_y', 'Tipo/Cor'], inplace=True)

df_merged['Descricao'] = df_merged['Descricao'].str.strip().apply(lambda x: x.ljust(30))
df_merged = df_merged.sort_values(by='Secao')
df_merged = df_merged[['Codigo', 'Deriv_x', 'Descricao', 'Rua', 'Secao', 'Andar', 'Quantidade']]
df_merged["QTD Entregue"] = " "
df_merged["Vol Entregue"] = " "
df_merged["Saldo Final"] = " "


df_merged.to_excel('tabela_imprimir.xlsx', index=False, engine='openpyxl')

wb = openpyxl.load_workbook('tabela_imprimir.xlsx')
ws = wb.active
ws.title = "Tabela"

# Adicione bordas de separação para linhas e colunas
borda = Border(left=Side(style='thin'), 
                     right=Side(style='thin'), 
                     top=Side(style='thin'), 
                     bottom=Side(style='thin'))

for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
    for cell in row:
        cell.border = borda

column_widths = {
    'A': 11,  # Código
    'B': 9,  # Deriv_x
    'C': 33,  # Descricao
    'D': 6,  # Rua
    'E': 6,  # Secao
    'F': 6,  # Andar
    'G': 13,  # Quantidade
    'H': 13,  # Quantidade Entregue
    'I': 13,  # Volumes Entregue
    'J': 13   # Saldo Final
}

for coluna, tamanho in column_widths.items():
    ws.column_dimensions[coluna].width = tamanho

for cell in ws[1]:
    cell.alignment = Alignment(horizontal='center')

# Salve o arquivo Excel
wb.save('tabela_imprimir.xlsx')

# Abra o arquivo Excel salvo
subprocess.Popen(['start', 'excel.exe', 'tabela_imprimir.xlsx'], shell=True)

print("Arquivo Excel aberto com sucesso!")









