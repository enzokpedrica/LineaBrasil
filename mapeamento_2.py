import pandas as pd
import openpyxl
from openpyxl.styles import Border, Side
import subprocess

arquivo = "Planilha.xlsx"

# Abrir o arquivo Excel
wb = openpyxl.load_workbook(arquivo)
ws = wb.active

# Criar uma borda sem estilo (sem borda)
sem_borda = Border(left=None, right=None, top=None, bottom=None)

# Remover bordas de todas as células
for row in ws.iter_rows():
    for cell in row:
        # Atribuir a borda sem estilo a todos os lados da célula
        cell.border = sem_borda

# Salvar o arquivo Excel sem bordas
wb.save(arquivo)

# Ler o arquivo Excel em um DataFrame do pandas
df = pd.read_excel(arquivo, engine="openpyxl")

# print(df[['Produto','Descrição Produto','Qtde Mov']])

file_map = "Lista_Map.CSV"
df_map = pd.read_csv(file_map, delimiter=';', encoding='latin1')

df_merged = pd.merge(df, df_map, left_on='Produto', right_on='Cod Produto', how="inner")

# Ordenar e selecionar as colunas necessárias
df_merged = df_merged.sort_values(by='Secao')
df_merged = df_merged[['Produto_x', 'Descrição Produto', 'Qtde Mov', 'Rua', 'Secao', 'Andar']]
df_merged = df_merged.drop_duplicates(subset=['Descrição Produto'])

df_merged.to_excel('tabela_imprimir.xlsx', index=False, engine='openpyxl')

wb = openpyxl.load_workbook('tabela_imprimir.xlsx')
ws = wb.active
ws.title = "Tabela"

# Salve o arquivo Excel
wb.save('tabela_imprimir.xlsx')

# Abra o arquivo Excel salvo
subprocess.Popen(['start', 'excel.exe', 'tabela_imprimir.xlsx'], shell=True)

